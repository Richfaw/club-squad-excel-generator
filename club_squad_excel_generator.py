import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import requests
from io import BytesIO
import tempfile

# Function to fetch logo
def fetch_logo(club_name):
    url = f"https://logo.clearbit.com/{club_name.replace(' ', '').lower()}.com"
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return BytesIO(response.content)
    except:
        pass
    return None

# Function to process squad data
def process_squad(raw_text, club_name):
    players = []
    lines = raw_text.strip().split("\n")
    for line in lines:
        parts = line.split("\t")
        if len(parts) < 2:
            parts = line.split()
        number = parts[0]
        name = parts[1]
        positions = parts[2:] if len(parts) > 2 else []
        primary = positions[0] if positions else ""
        secondary = ", ".join(positions[1:]) if len(positions) > 1 else ""
        players.append((number, name, primary, secondary))

    wb = Workbook()
    ws = wb.active
    ws.title = "Squad"

    # Add title
    ws.merge_cells('A1:O1')
    ws['A1'] = f"{club_name} Squad List"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Number", "Name", "GK", "CB", "LB", "RB", "DM", "CM", "RM", "LM", "AM", "LW", "RW", "SS", "CF"]
    ws.append(headers)

    fills = {
        "GK": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "CB": PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid"),
        "LB": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "RB": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "LM": PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
        "LW": PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
        "DM": PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid"),
        "CM": PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid"),
        "AM": PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid"),
        "RM": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
        "RW": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
        "SS": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
        "CF": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
    }

    # Insert player data
    for idx, (number, name, primary, secondary) in enumerate(players, start=3):
        row_data = {pos: "" for pos in headers}
        row_data["Number"] = number
        row_data["Name"] = name

        if primary in headers:
            row_data[primary] = primary
        if secondary:
            for pos in secondary.split(", "):
                if pos in headers:
                    row_data[pos] = pos

        row = [row_data[col] for col in headers]
        ws.append(row)

        for col_idx, pos in enumerate(headers[2:], start=3):
            if row_data[pos] != "":
                cell = ws.cell(row=idx, column=col_idx)
                if pos in fills:
                    cell.fill = fills[pos]

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    logo_data = fetch_logo(club_name)
    if logo_data:
        img = Image(logo_data)
        img.height = 80
        img.width = 80
        ws.add_image(img, 'P1')

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    return temp_file.name

st.title("⚽ Club Squad Excel Generator")

raw_input_text = st.text_area("Paste your squad list here:", height=300)
club = st.text_input("Enter club name (e.g., Everton FC):")

if st.button("Generate Excel"):
    if raw_input_text and club:
        file_path = process_squad(raw_input_text, club)
        with open(file_path, "rb") as f:
            st.download_button("Download Squad Excel", f, file_name=f"{club.replace(' ', '_').lower()}_squad.xlsx")
    else:
        st.error("Please provide both squad list and club name.")